"""DGFiP Excel parser (jalon 6.5).

The pipeline parses the real DGFiP ISF/IFI workbooks instead of only the curated
CSV / pre-filled points. The IFI page publishes the *same* population sliced
three ways (déciles de patrimoine, déciles de RFR, tranches de taux marginal);
each slice namespaces its index into `groupe` so they never collide on
HIST_KEYS. These tests build *synthetic* workbooks mirroring the real layouts
(no binary fixtures committed) and assert:

- each IFI breakdown -> its own `groupe` prefix, the right indicateurs, and
  amounts/counts converted from milliers to base units (×1000);
- the three breakdowns coexist in one directory with zero HIST_KEYS collisions;
- ISF (`nombres` sheet) -> `nb_foyers` per year, Convention `total`;
- the dispatcher ignores commune-level workbooks and non-Excel files;
- `load_dgfip` parses a directory and falls back to DGFIP_POINTS (issue #12).
"""

from __future__ import annotations

import sys
from pathlib import Path

import openpyxl

# The parser + loader live in the pipeline package at the repo root.
PIPELINE_DIR = Path(__file__).resolve().parents[2] / "pipeline"
if str(PIPELINE_DIR) not in sys.path:
    sys.path.insert(0, str(PIPELINE_DIR))

import build_dataset  # noqa: E402
import dgfip_parse  # noqa: E402

HIST_KEYS = ["annee", "source", "concept_patrimoine", "unite", "groupe", "indicateur"]


def _sheet_with_header(wb, name: str, header: list, data_rows: list, note_row: bool = True):
    ws = wb.create_sheet(name)
    ws.append([None] * len(header))  # row 0: blank/title
    ws.append([None] * len(header))  # row 1: unit caption
    ws.append(header)  # row 2: header
    for row in data_rows:
        ws.append(row)
    if note_row:
        ws.append(["Source : Ministère de l'Économie", *[None] * (len(header) - 1)])


def _write_ifi_patrimoine(path: Path, years: list[int]) -> None:
    """node/25582: `a{year}` sheets, déciles de patrimoine (impôt + bornes)."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    header = [
        "NUMÉRO DE DÉCILE DE PATRIMOINE TAXABLE",
        "BORNE INFÉRIEURE DU DÉCILE DE PATRIMOINE",
        "BORNE SUPÉRIEURE DU DÉCILE DE PATRIMOINE",
        "IMPÔT NET MOYEN",
        "TOTAL DES DONS",
        "DON MOYEN",
    ]
    for y in years:
        rows = [
            [d, 1000 + (d - 1) * 100, ("." if d == 10 else 1000 + d * 100), d * 2.5, 9.0, 1.0]
            for d in range(1, 11)
        ]
        _sheet_with_header(wb, f"a{y}", header, rows)
    wb.save(path)


def _write_ifi_rfr(path: Path, years: list[int]) -> None:
    """node/25583: `ifi_{year}` sheets, déciles de RFR (+ patrimoine moyen)."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    header = [
        "NUMÉRO DE DÉCILE DE RFR DES REDEVABLES",
        "BORNE INFÉRIEURE DU DÉCILE DE RFR",
        "BORNE SUPÉRIEURE DU DÉCILE DE RFR",
        "PATRIMOINE NET TAXABLE MOYEN",
        "IMPÔT NET MOYEN",
        "TOTAL DES DONS",
        "DON MOYEN",
    ]
    for y in years:
        rows = [
            [
                d,
                30 + (d - 1) * 5,
                ("." if d == 10 else 30 + d * 5),
                1500 + d * 10,
                d * 2.5,
                9.0,
                1.0,
            ]
            for d in range(1, 11)
        ]
        _sheet_with_header(wb, f"ifi_{y}", header, rows)
    wb.save(path)


def _write_ifi_tranche(path: Path, years: list[int]) -> None:
    """node/25584: `a{year}` sheets, tranches de taux marginal (+ nb redevables)."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    header = [
        "NUMÉRO DE LA TRANCHE MARGINALE DE TAXATION",
        "BORNE INFÉRIEURE DE LA TRANCHE MARGINALE",
        "BORNE SUPÉRIEURE DE LA TRANCHE MARGINALE",
        "TAUX DE LA TRANCHE MARGINALE DE TAXATION",
        "NOMBRE DE REDEVABLES",
        "PATRIMOINE NET TAXABLE MOYEN",
        "IMPÔT NET MOYEN",
    ]
    for y in years:
        # Tranches are numbered from 2 (tranche 1 is below the taxable threshold).
        rows = [
            [
                t,
                1300 * (t - 1),
                ("." if t == 4 else 1300 * t),
                0.5 + t * 0.1,
                100 * t,
                1500 + t * 10,
                t * 5.0,
            ]
            for t in range(2, 5)
        ]
        _sheet_with_header(wb, f"a{y}", header, rows)
    wb.save(path)


def _write_isf_nombres(path: Path, years: list[int]) -> None:
    """`isf_montants_declares_nombres_*.xls`: `nombres` + `montants` sheets."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("nombres")
    ws.append(["Déclarations ISF", *[None] * len(years)])  # row 0: title
    ws.append(["Cases", "Libellé des cases", *years])  # row 1: year header
    ws.append([None, "nombre de déclarations", *[200000 + y for y in years]])  # row 2
    ws.append(["AB", "Immeubles bâtis", *[1 for _ in years]])
    wb.create_sheet("montants")  # presence only; unused by the parser
    wb.save(path)


def test_parse_ifi_patrimoine(tmp_path: Path) -> None:
    f = tmp_path / "patrimoine.xlsx"
    _write_ifi_patrimoine(f, [2018, 2019])
    rows = dgfip_parse.parse_ifi_breakdown(f)

    # 2 years x (10 impot_moyen + 9 seuil; open top décile has no borne) = 38
    assert len(rows) == 38
    assert all(r["groupe"].startswith("decile_patrimoine_") for r in rows)
    assert {r["concept_patrimoine"] for r in rows} == {"immobilier"}
    assert {r["indicateur"] for r in rows} == {"impot_moyen", "seuil"}

    d1 = next(
        r
        for r in rows
        if r["annee"] == 2018
        and r["groupe"] == "decile_patrimoine_1"
        and r["indicateur"] == "impot_moyen"
    )
    assert d1["valeur"] == 2500.0  # 2.5 milliers -> 2500 euros
    assert d1["unite_valeur"] == "euros"
    assert d1["millesime_source"] == "DGFiP IFI patrimoine 2019"
    assert not [
        r for r in rows if r["groupe"] == "decile_patrimoine_10" and r["indicateur"] == "seuil"
    ]


def test_parse_ifi_rfr_includes_patrimoine_moyen(tmp_path: Path) -> None:
    f = tmp_path / "rfr.xlsx"
    _write_ifi_rfr(f, [2020])
    rows = dgfip_parse.parse_ifi_breakdown(f)

    assert all(r["groupe"].startswith("decile_rfr_") for r in rows)
    assert {r["indicateur"] for r in rows} == {"patrimoine_moyen", "impot_moyen", "seuil"}
    pm = next(
        r for r in rows if r["groupe"] == "decile_rfr_1" and r["indicateur"] == "patrimoine_moyen"
    )
    assert pm["valeur"] == (1500 + 10) * 1000  # 1510 milliers -> euros
    assert all(r["millesime_source"] == "DGFiP IFI RFR 2020" for r in rows)


def test_parse_ifi_tranche_includes_nb_foyers_and_taux_note(tmp_path: Path) -> None:
    f = tmp_path / "tranche.xlsx"
    _write_ifi_tranche(f, [2018])
    rows = dgfip_parse.parse_ifi_breakdown(f)

    assert all(r["groupe"].startswith("tranche_marginale_") for r in rows)
    assert {r["indicateur"] for r in rows} == {
        "nb_foyers",
        "patrimoine_moyen",
        "impot_moyen",
        "seuil",
    }
    nb = next(
        r for r in rows if r["groupe"] == "tranche_marginale_2" and r["indicateur"] == "nb_foyers"
    )
    assert nb["valeur"] == 200 * 1000  # 100*2 milliers -> 200000 effectif
    assert nb["unite_valeur"] == "effectif"
    # The marginal rate is recorded in notes, never multiplied by 1000.
    assert "taux marginal 0.7 %" in nb["notes"]
    assert all(r["millesime_source"] == "DGFiP IFI taux marginal 2018" for r in rows)


def test_three_breakdowns_do_not_collide(tmp_path: Path) -> None:
    _write_ifi_patrimoine(tmp_path / "p.xlsx", [2018, 2019])
    _write_ifi_rfr(tmp_path / "r.xlsx", [2018, 2019])
    _write_ifi_tranche(tmp_path / "t.xlsx", [2018, 2019])

    df = build_dataset.load_dgfip(tmp_path, "2026-06-15", "DGFiP test")

    prefixes = {g.rsplit("_", 1)[0] for g in df["groupe"]}
    assert prefixes == {"decile_patrimoine", "decile_rfr", "tranche_marginale"}
    # No two rows share the full historisation key within one millésime.
    dup = df.duplicated(subset=HIST_KEYS + ["millesime_source"]).sum()
    assert dup == 0


def test_parse_isf_montants_nb_foyers(tmp_path: Path) -> None:
    f = tmp_path / "isf.xlsx"
    _write_isf_nombres(f, [1999, 2000, 2017])
    rows = dgfip_parse.parse_isf_montants(f)

    assert len(rows) == 3
    assert all(r["concept_patrimoine"] == "total" for r in rows)
    assert all(r["indicateur"] == "nb_foyers" and r["unite_valeur"] == "effectif" for r in rows)
    assert {r["annee"]: r["valeur"] for r in rows} == {1999: 201999, 2000: 202000, 2017: 202017}
    assert all(r["millesime_source"] == "DGFiP ISF 2017" for r in rows)


def test_dispatcher_recognises_and_ignores(tmp_path: Path) -> None:
    _write_ifi_patrimoine(tmp_path / "p.xlsx", [2018])
    assert len(dgfip_parse.parse_dgfip_excel(tmp_path / "p.xlsx")) == 19
    _write_ifi_rfr(tmp_path / "r.xlsx", [2018])  # ifi_{year} sheets also recognised
    assert dgfip_parse.parse_dgfip_excel(tmp_path / "r.xlsx")
    _write_isf_nombres(tmp_path / "isf.xlsx", [2017])
    assert len(dgfip_parse.parse_dgfip_excel(tmp_path / "isf.xlsx")) == 1

    # Commune-level workbook: a single sheet that matches neither layout.
    com = tmp_path / "ificom2024.xlsx"
    wb = openpyxl.Workbook()
    wb.active.title = "IFI 2024"
    wb.active.append(["Ville", "Redevables"])
    wb.save(com)
    assert dgfip_parse.parse_dgfip_excel(com) == []

    # A non-Excel file does not raise.
    txt = tmp_path / "notes.csv"
    txt.write_text("annee,valeur\n2020,1\n")
    assert dgfip_parse.parse_dgfip_excel(txt) == []


def test_num_handles_dgfip_sentinels() -> None:
    assert dgfip_parse._num(".") is None
    assert dgfip_parse._num("") is None
    assert dgfip_parse._num(None) is None
    assert dgfip_parse._num("1 674") == 1674.0  # thin-space grouped
    assert dgfip_parse._num("5,4") == 5.4  # comma decimal
    assert dgfip_parse._num(43.9) == 43.9


def test_load_dgfip_parses_directory_and_annotates_rupture(tmp_path: Path) -> None:
    _write_ifi_patrimoine(tmp_path / "ifi.xlsx", [2018])
    _write_isf_nombres(tmp_path / "isf.xlsx", [2017])

    df = build_dataset.load_dgfip(tmp_path, "2026-06-15", "DGFiP test")

    assert set(df["source"]) == {"DGFiP"}
    assert len(df) == 20  # IFI 2018 (19) + ISF 2017 (1)
    assert {"immobilier", "total"} <= set(df["concept_patrimoine"])
    assert df["notes"].str.contains("RUPTURE ISF->IFI 2018").any()
    assert (df["date_extraction"] == "2026-06-15").all()


def test_load_dgfip_falls_back_to_points(tmp_path: Path) -> None:
    """A directory with no recognisable workbook -> pre-filled DGFIP_POINTS."""
    (tmp_path / "ignored.txt").write_text("nothing here")
    df = build_dataset.load_dgfip(tmp_path, "2026-06-15", "DGFiP test")
    assert len(df) == len(build_dataset.DGFIP_POINTS)
    assert set(df["source"]) == {"DGFiP"}
